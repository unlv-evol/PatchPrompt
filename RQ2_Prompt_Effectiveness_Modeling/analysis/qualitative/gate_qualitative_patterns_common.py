"""Shared utilities for gate-specific qualitative evidence tables.

These helpers build the gate-specific qualitative evidence artifacts used in the
replication package.  They deliberately operate from the canonical final dataset
(`Dataset_Construction/processed_data/final_analysis_dataset.csv`) so that the curated evidence
records can be traced back to the exact downstream dataset used in the paper.

Generated artifacts per gate:
- CSV table for machine-readable inspection
- XLSX table for spreadsheet review
- PDF table with readable black body text
- full canonical case-record CSV for traceability
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:  # pragma: no cover
    colors = None


ROOT = Path(__file__).resolve().parents[2]
DATASET_PATH = ROOT.parent / "Dataset_Construction" / "processed_data" / "final_analysis_dataset.csv"
RESULTS_ROOT = ROOT / "results" / "qualitative"


def load_final_dataset() -> pd.DataFrame:
    """Load the canonical final analysis dataset and standardize lookup fields."""
    df = pd.read_csv(DATASET_PATH)
    df["Case ID"] = df["Case ID"].astype(str)
    return df


def _case_lookup(df: pd.DataFrame, case_id: str) -> pd.Series:
    rows = df.loc[df["Case ID"] == case_id]
    if rows.empty:
        raise ValueError(f"Case ID {case_id!r} not found in {DATASET_PATH}")
    return rows.iloc[0]


def _score(value) -> int:
    """Convert rubric score values from float-like CSV values to integer labels."""
    if pd.isna(value):
        return 0
    return int(round(float(value)))


def _ensure_english_excerpt(text: str, gate: str, case_id: str) -> str:
    """Require English-only prompt excerpts for evaluator-facing artifacts.

    We enforce ASCII here to prevent multilingual excerpts from leaking into the
    generated CSV/XLSX/PDF bundles. Source case specs should provide translated
    English excerpts before generation.
    """
    if any(ord(ch) > 127 for ch in text):
        raise ValueError(
            f"Non-ASCII Prompt Excerpt found for {gate} case {case_id}. "
            "Provide an English translation in the case spec."
        )
    return text


def build_pattern_table(specs: Iterable[dict], gate: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build a curated qualitative-pattern table and full-record trace table.

    Parameters
    ----------
    specs:
        Curated evidence-case specifications containing case IDs, prompt excerpts,
        structural pattern labels, and concise interpretations.
    gate:
        Gate identifier such as ``gate0``, ``gate1``, or ``gate2``.

    Returns
    -------
    pattern_df, full_records_df:
        The spreadsheet-facing curated table and the corresponding full canonical
        dataset records for traceability.
    """
    source = load_final_dataset()
    rows = []
    full_records = []

    for item in specs:
        case = _case_lookup(source, item["Case ID"])
        full_records.append(case.to_dict())

        base = {
            "Case ID": case["Case ID"],
            "Outcome Class": case["Outcome_Class"],
            "Context (C)": _score(case["Context"]),
            "Specificity (S)": _score(case["Specificity"]),
            "Verification (V)": _score(case["Verification"]),
            "Structural Pattern": item["Structural Pattern"],
            "Prompt Excerpt": _ensure_english_excerpt(item["Prompt Excerpt"], gate, case["Case ID"]),
            "Why Representative": item["Why Representative"],
            "Notes": item.get("Notes", ""),
            "PR Language": case.get("PR_Language", ""),
            "PR Link": case.get("PR_Link", ""),
            "Conversation Link": case.get("Conversation_Link", ""),
        }

        if gate == "gate0":
            base = {
                "Case ID": base["Case ID"],
                "Outcome Class (PA / PN / NE)": base["Outcome Class"],
                "Context Score (C)": base["Context (C)"],
                "Specificity Score (S)": base["Specificity (S)"],
                "Verification Score (V)": base["Verification (V)"],
                "Structural Pattern": base["Structural Pattern"],
                "Prompt Excerpt": base["Prompt Excerpt"],
                "Why Representative": base["Why Representative"],
                "Generated Actionable Code?": item["Generated Actionable Code?"],
                "Notes": base["Notes"],
                "PR Language": base["PR Language"],
                "PR Link": base["PR Link"],
                "Conversation Link": base["Conversation Link"],
            }
        elif gate == "gate1":
            base["Generated Code?"] = item["Generated Code?"]
            base["Adopted?"] = item["Adopted?"]
            order = [
                "Case ID", "Outcome Class", "Context (C)", "Specificity (S)",
                "Verification (V)", "Structural Pattern", "Prompt Excerpt",
                "Why Representative", "Generated Code?", "Adopted?", "Notes",
                "PR Language", "PR Link", "Conversation Link",
            ]
            base = {k: base[k] for k in order}
        elif gate == "gate2":
            base = {
                "Case ID": base["Case ID"],
                "Integration Fraction": round(float(case.get("Fraction_Adopted", 0)), 2),
                "Context (C)": base["Context (C)"],
                "Specificity (S)": base["Specificity (S)"],
                "Verification (V)": base["Verification (V)"],
                "Structural Pattern": base["Structural Pattern"],
                "Prompt Excerpt": base["Prompt Excerpt"],
                "Why Representative": base["Why Representative"],
                "Integration Depth Category": item["Integration Depth Category"],
                "Notes": base["Notes"],
                "Outcome Class": base["Outcome Class"],
                "PR Language": base["PR Language"],
                "PR Link": base["PR Link"],
                "Conversation Link": base["Conversation Link"],
            }
        else:
            raise ValueError(f"Unknown gate: {gate}")

        rows.append(base)

    return pd.DataFrame(rows), pd.DataFrame(full_records)


def write_csv_xlsx_pdf(pattern_df: pd.DataFrame, full_df: pd.DataFrame, gate: str, title: str) -> None:
    """Write gate-specific qualitative evidence artifacts under results."""
    result_dir = RESULTS_ROOT / gate
    result_dir.mkdir(parents=True, exist_ok=True)

    csv_path = result_dir / f"{gate}_qualitative_patterns.csv"
    xlsx_path = result_dir / f"{gate}_qualitative_patterns.xlsx"
    pdf_path = result_dir / f"{gate}_qualitative_patterns.pdf"
    full_path = result_dir / f"{gate}_full_case_records.csv"

    pattern_df.to_csv(csv_path, index=False)
    full_df.to_csv(full_path, index=False)
    pattern_df.to_excel(xlsx_path, index=False)

    _style_xlsx(xlsx_path)
    _write_pdf(pattern_df, pdf_path, title)

    _update_examples_manifest(pattern_df, gate)


def _update_examples_manifest(pattern_df: pd.DataFrame, gate: str) -> None:
    """Refresh the cross-gate qualitative examples manifest from current outputs."""
    outcome_col = "Outcome Class"
    if outcome_col not in pattern_df.columns and "Outcome Class (PA / PN / NE)" in pattern_df.columns:
        outcome_col = "Outcome Class (PA / PN / NE)"

    required = ["Case ID", outcome_col, "Structural Pattern"]
    missing = [c for c in required if c not in pattern_df.columns]
    if missing:
        raise ValueError(f"Cannot update qualitative examples manifest; missing columns: {missing}")

    gate_label = gate.upper()
    updated = pattern_df[["Case ID", outcome_col, "Structural Pattern"]].copy()
    updated = updated.rename(columns={outcome_col: "Outcome Class"})
    updated.insert(0, "Gate", gate_label)

    manifest_path = RESULTS_ROOT / "qualitative_examples_manifest.csv"
    if manifest_path.exists():
        existing = pd.read_csv(manifest_path)
        for col in ["Gate", "Case ID", "Outcome Class", "Structural Pattern"]:
            if col not in existing.columns:
                existing[col] = ""
        existing = existing[["Gate", "Case ID", "Outcome Class", "Structural Pattern"]]
        existing = existing[existing["Gate"] != gate_label]
        merged = pd.concat([existing, updated], ignore_index=True)
    else:
        merged = updated

    gate_order = {"GATE0": 0, "GATE1": 1, "GATE2": 2}
    merged["__gate_order"] = merged["Gate"].map(gate_order).fillna(99)
    merged = merged.sort_values(["__gate_order", "Case ID"], kind="stable").drop(columns=["__gate_order"])
    merged.to_csv(manifest_path, index=False)


def _style_xlsx(path: Path) -> None:
    """Apply light spreadsheet formatting for reviewer readability."""
    if load_workbook is None:
        return
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="000000")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    widths = {
        "A": 12, "B": 18, "C": 12, "D": 14, "E": 14, "F": 30,
        "G": 60, "H": 60, "I": 18, "J": 24, "K": 18, "L": 20,
        "M": 45, "N": 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 72
    wb.save(path)


def _register_unicode_fonts() -> tuple[str, str]:
    """Register Unicode-capable fonts for multilingual prompt excerpts.

    ReportLab's built-in Helvetica font does not contain Japanese/CJK glyphs,
    which causes non-English prompt excerpts to render as square boxes in the
    PDF tables.  We therefore use system-installed Unicode fonts when available.
    No font files are redistributed with the package; the code falls back to
    Helvetica only when no suitable system font is available.
    """
    if colors is None:
        return "Helvetica", "Helvetica-Bold"

    regular_candidates = [
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    bold_candidates = [
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]

    regular_name = "Helvetica"
    bold_name = "Helvetica-Bold"

    for candidate in regular_candidates:
        font_path = Path(candidate)
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("QualUnicode", str(font_path)))
                regular_name = "QualUnicode"
                break
            except Exception:
                continue

    for candidate in bold_candidates:
        font_path = Path(candidate)
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("QualUnicodeBold", str(font_path)))
                bold_name = "QualUnicodeBold"
                break
            except Exception:
                continue

    return regular_name, bold_name


def _write_pdf(df: pd.DataFrame, path: Path, title: str) -> None:
    """Render a readable multilingual PDF table with black body text.

    Ensure PDFs render Japanese/CJK prompt excerpts correctly by using a
    Unicode-capable system font and CJK-aware wrapping.  The default Helvetica
    font lacks those glyphs, which causes non-English text to appear as square
    boxes.  This version registers a suitable system font when available so
    non-English text appears as it does in the CSV/XLSX files.
    """
    if colors is None:
        path.write_text("PDF export skipped because reportlab is not installed.\n", encoding="utf-8")
        return

    body_font, bold_font = _register_unicode_fonts()

    # Keep CSV/XLSX fully traceable, but omit long URL columns in the PDF
    # preview so the evidence table remains readable and does not clip.
    df = df.drop(columns=["PR Link", "Conversation Link"], errors="ignore")

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A3),
        rightMargin=0.20 * inch,
        leftMargin=0.20 * inch,
        topMargin=0.22 * inch,
        bottomMargin=0.22 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleBlack", parent=styles["Title"], textColor=colors.black,
        fontName=bold_font, fontSize=18, leading=22, alignment=1,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "HeaderWhite", parent=styles["Normal"], textColor=colors.white,
        fontName=bold_font, fontSize=7.0, leading=8.0,
        wordWrap="CJK",
    )
    cell_style = ParagraphStyle(
        "CellBlack", parent=styles["Normal"], textColor=colors.black,
        fontName=body_font, fontSize=5.8, leading=7.0,
        wordWrap="CJK", splitLongWords=True,
    )

    def para(value, style):
        safe = "" if pd.isna(value) else str(value)
        safe = safe.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(safe, style)

    data = [[para(col, header_style) for col in df.columns]]
    for _, row in df.iterrows():
        data.append([para(row[col], cell_style) for col in df.columns])

    ncols = len(df.columns)
    if ncols == 11:  # gate0 after dropping URL columns
        widths = [0.60, 0.90, 0.60, 0.70, 0.70, 1.50, 3.90, 3.70, 0.90, 1.40, 0.90]
    elif ncols == 12:  # gate1/gate2 after dropping URL columns
        widths = [0.55, 0.70, 0.50, 0.55, 0.55, 1.35, 3.50, 3.30, 0.75, 0.75, 1.00, 0.75]
    elif ncols == 13:  # legacy gate0 with URL columns
        widths = [0.55, 0.80, 0.55, 0.60, 0.60, 1.20, 2.40, 2.30, 0.70, 1.10, 0.70, 1.40, 1.40]
    elif ncols == 14:  # legacy gate1/gate2 with URL columns
        widths = [0.50, 0.65, 0.50, 0.55, 0.55, 1.25, 2.25, 2.10, 0.75, 0.75, 1.00, 0.65, 1.25, 1.25]
    else:
        widths = [15.0 / ncols] * ncols
    col_widths = [w * inch for w in widths]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("FONTNAME", (0, 1), (-1, -1), body_font),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7B7B7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6FA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    doc.build([Paragraph(title, title_style), Spacer(1, 0.12 * inch), table])


def write_gate_readme(gate: str, text: str) -> None:
    root = RESULTS_ROOT / gate
    root.mkdir(parents=True, exist_ok=True)
    (root / "README.txt").write_text(text.strip() + "\n", encoding="utf-8")
